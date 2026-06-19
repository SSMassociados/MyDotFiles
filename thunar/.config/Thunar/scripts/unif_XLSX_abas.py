#!/usr/bin/env python3
import pandas as pd
import os
import sys
import subprocess
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def auto_ajustar_largura_colunas(planilha):
    """
    Ajusta automaticamente a largura das colunas conforme o conteúdo
    """
    for col in planilha.columns:
        max_length = 0
        coluna = col[0].column_letter  # Get the column name
        
        for cell in col:
            try:
                # Calcular comprimento do conteúdo
                if cell.value:
                    comprimento = len(str(cell.value))
                    if comprimento > max_length:
                        max_length = comprimento
            except:
                pass
        
        # Adicionar margem de segurança
        largura_ajustada = (max_length + 2)
        # Limitar largura máxima (opcional, evita colunas muito largas)
        if largura_ajustada > 50:
            largura_ajustada = 50
        
        planilha.column_dimensions[coluna].width = largura_ajustada

# Obter arquivos selecionados
arquivos = sys.argv[1:]
if not arquivos:
    sys.exit(1)

# Criar subdiretório para arquivos unificados
subpasta = "XLSX_Unificado"
if not os.path.exists(subpasta):
    os.makedirs(subpasta)
    print(f"Diretório '{subpasta}' criado.")

# Gerar nome do arquivo de saída com timestamp
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
nome_arquivo = f"unificado_{timestamp}.xlsx"
saida = os.path.join(subpasta, nome_arquivo)

try:
    # Primeiro: criar o arquivo Excel com os dados
    with pd.ExcelWriter(saida, engine='openpyxl') as writer:
        for i, arquivo in enumerate(arquivos, 1):
            if arquivo.lower().endswith(('.xlsx', '.xls', '.xlsm')):
                try:
                    # Obter nome base para a guia
                    nome_base = os.path.splitext(os.path.basename(arquivo))[0]
                    
                    # Sanitizar nome da guia (remover caracteres inválidos)
                    nome_guia = ''.join(c for c in nome_base if c not in r'[]:*?/\\')[:31]
                    
                    # Ler e escrever dados
                    df = pd.read_excel(arquivo)
                    df.to_excel(writer, sheet_name=nome_guia, index=False)
                    
                    print(f"Processado ({i}/{len(arquivos)}): {os.path.basename(arquivo)}")
                    
                except Exception as e:
                    print(f"Erro ao processar {os.path.basename(arquivo)}: {str(e)}")
                    continue
    
    # Segundo: carregar o arquivo criado e ajustar as colunas
    print("Ajustando largura das colunas...")
    workbook = load_workbook(saida)
    
    for sheet_name in workbook.sheetnames:
        planilha = workbook[sheet_name]
        auto_ajustar_largura_colunas(planilha)
        print(f"  Colunas ajustadas: {sheet_name}")
    
    # Salvar as modificações
    workbook.save(saida)
    
    # Mensagem de sucesso no terminal
    print(f"✅ Arquivo criado: {saida}")
    
    # Notificação desktop
    try:
        subprocess.run([
            'notify-send', '-u', 'normal', '-t', '3000',
            'Unificação Concluída', 
            f'Arquivo salvo: {nome_arquivo}\nColunas auto-ajustadas!'
        ])
    except Exception:
        print("⚠️  notify-send não disponível")
    
    # Som de conclusão
    try:
        subprocess.run([
            'paplay', '/usr/share/sounds/freedesktop/stereo/complete.oga'
        ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        print("⚠️  Som de conclusão não disponível")
        
except Exception as e:
    # Notificação de erro
    error_msg = f"Erro na unificação: {str(e)}"
    print(f"❌ {error_msg}")
    
    try:
        subprocess.run([
            'notify-send', '-u', 'critical', '-t', '5000',
            'Erro na Unificação', 
            error_msg
        ])
    except Exception:
        pass
    
    sys.exit(1)
